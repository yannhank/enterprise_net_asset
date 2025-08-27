import base64
import time,os
import socket
from openpyxl import load_workbook
import requests
import json,secrets,string
import argparse
import pymysql

db_user = "qiyelist"
db_name = "qiyelist"
db_pass = "123456"
db = pymysql.connect(host='localhost',user=db_user,password=db_pass,database=db_name,charset='utf8')
table_name = ""
file_path = ""

def start():
    path = f"{file_path}/"
    flist = os.listdir(path)
    print(flist)
    for fname in flist:
        ff = os.path.join(path, fname)
        res = readXlsx(ff,4)
        os.remove(ff)
    print("")

def getCode(length=8):
    chars = string.ascii_letters + string.digits
    return ''.join(secrets.choice(chars) for _ in range(length))

def readXlsx(csv,line):
    wb = load_workbook(f'{csv}')
    sheet = wb.active
    companys = []
    for row in sheet.iter_rows(min_row=line, values_only=True):
        code = getCode()
        data = [code,row[0],row[1],row[2],row[3],row[5],row[9],row[10],row[11],row[12],row[13],row[15],row[20],row[21],row[23],row[25]]
        #替换数组里的-符号为空
        res = [item if item != '-' else '' for item in data]
        option(res)

def option(data):
    qy_name = data[0]
    qy_domain = data[15]
    chk = chkData(qy_name,qy_domain)
    if chk == True:
        pass
    else:
        if data[2]=="开业":
            insert(data)

def chkData(qy_name,qy_domain):
    s = f"select * from qiye_{table_name} where qy_name='{qy_name}' and qy_domain='{qy_domain}'"
    c = db.cursor()
    c.execute(s)
    d = c.fetchone()
    c.close()
    if d:
        print("-",end="",flush=True)
        return True
    else:
        return False

def insert(data):
    domain = ""
    a = data[15].split("//")
    if len(a)>1:
        b = a[1].split("/")
        domain = b[0]
    else:
        bb = a[0].split("/")
        domain = bb[0]
    tt = domain.split("www.")
    if len(tt)>1:
        domain = tt[1]
    else:
        domain = tt[0]
    s = (f"INSERT INTO `qiye_{table_name}` (`code`, `qy_name`, `qy_status`, `qy_fanren`, `qy_ziben`,`qy_date`, `city`, "
         f"`quxian`,`tel1`,`tel2`,`qy_email`,`qy_shuihao`,`qy_type`,`qy_hangye`,`qy_address`,`qy_domain`,`domain_status`,`company_status`) VALUES "
         f"('{data[0]}', '{data[1]}', '{data[2]}', '{data[3]}', '{data[4]}', '{data[5]}', '{data[6]}', '{data[7]}', "
         f"'{data[8]}', '{data[9]}', '{data[10]}', '{data[11]}', '{data[12]}', '{data[13]}','{data[14]}','{domain}', '0','0')")
    #print(s)
    try:
        c = db.cursor()
        c.execute(s)
        db.commit()
        c.close()
        print("+",end="",flush=True)
    except Exception as e:
        print(s)

if __name__ == '__main__':
    parser = argparse.ArgumentParser()
    parser.add_argument('-t', type=str, help='企业列表')
    parser.add_argument('-p', type=str, help='xlsx文件路径')
    args = parser.parse_args()
    #企业列表
    table_name = args.t
    # xlsx文件路径
    file_path = args.p
    start()